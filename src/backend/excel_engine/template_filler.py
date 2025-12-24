"""
Template Filler - AESIA Excel Template Processing
Fills official AESIA checklist templates with assessment data while preserving formulas and styling.
"""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from typing import Dict, List, Optional, Any
from pathlib import Path
import os

# Template file mapping: requirement_code -> filename
TEMPLATE_MAPPING = {
    'QUALITY_MGMT': 'Gestión de Calidad_Checklist.xlsx',
    'RISK_MGMT': 'Gestión de riesgos_Checklist.xlsx',
    'HUMAN_OVERSIGHT': 'Supervisión Humana_Checklist.xlsx',
    'DATA_GOVERNANCE': 'Gobernanza del dato_Checklist.xlsx',
    'TRANSPARENCY': 'Transparencia_Checklist.xlsx',
    'ACCURACY': 'Precision_Checklist.xlsx',
    'ROBUSTNESS': 'Solidez_Checklist.xlsx',
    'CYBERSECURITY': 'Ciberseguridad_Checklist.xlsx',
    'LOGGING': 'Registros_Checklist.xlsx',
    'TECHNICAL_DOC': 'Documentación tecnica_Checklist.xlsx',
    'POST_MARKET': 'Vigilancia Poscomercializacion_Checklist.xlsx',
    'INCIDENT_MGMT': 'Gestión de incidentes_Checklist.xlsx',
}

# Expected sheet names in AESIA templates (in order)
EXPECTED_SHEETS = [
    'Portada',
    'Intro',
    'Artículo RIA',
    'Medidas Guías',
    'Relación MG-Apart',
    'Autoeval MG',
    'Medidas Adicionales',
    'Relación MA-Apart',
    'Autoeval MA',
]

# Column headers to detect (case-insensitive matching)
COLUMN_MATCHERS = {
    'difficulty': ['nivel de dificultad', 'dificultad percibida', 'dificultad'],
    'maturity': ['nivel de madurez', 'madurez', 'nivel madurez'],
    'mg_id': ['mg', 'medida guía', 'id medida', 'código mg'],
    'subpart_id': ['apartado', 'subapartado', 'art.', 'artículo'],
    'description': ['descripción', 'descripcion', 'detalle'],
    'file_name': ['archivo', 'nombre archivo', 'documento'],
}


class TemplateFiller:
    """Fills AESIA Excel templates with assessment data."""
    
    def __init__(self, templates_dir: str = None):
        """
        Initialize template filler.
        
        Args:
            templates_dir: Path to directory containing template files.
                          If None, uses ./templates/ relative to this file.
        """
        if templates_dir is None:
            templates_dir = os.path.join(os.path.dirname(__file__), 'templates')
        self.templates_dir = Path(templates_dir)
    
    def get_template_path(self, requirement_code: str) -> Optional[Path]:
        """Get path to template file for given requirement code."""
        filename = TEMPLATE_MAPPING.get(requirement_code)
        if not filename:
            return None
        path = self.templates_dir / filename
        return path if path.exists() else None
    
    def fill_template(
        self,
        requirement_code: str,
        assessments_mg: List[Dict[str, Any]],
        measures_additional: List[Dict[str, Any]] = None,
        assessments_ma: List[Dict[str, Any]] = None,
        ma_to_subpart: List[Dict[str, str]] = None,
        application_info: Dict[str, Any] = None,
    ) -> bytes:
        """
        Fill a template with assessment data and return as bytes.
        
        Args:
            requirement_code: The requirement code (e.g., 'QUALITY_MGMT')
            assessments_mg: List of MG assessments with mg_id, subpart_id, difficulty, maturity
            measures_additional: List of additional measures
            assessments_ma: List of MA assessments
            ma_to_subpart: List of MA to subpart relationships
            application_info: Application metadata (name, sector, etc.)
        
        Returns:
            Excel file as bytes
        """
        template_path = self.get_template_path(requirement_code)
        if not template_path:
            raise ValueError(f"No template found for requirement: {requirement_code}")
        
        # Load workbook preserving formulas
        wb = load_workbook(template_path)
        
        # Fill each relevant sheet
        self._fill_autoeval_mg(wb, assessments_mg)
        
        if measures_additional:
            self._fill_measures_additional(wb, measures_additional)
        
        if ma_to_subpart:
            self._fill_relation_ma(wb, ma_to_subpart, measures_additional or [])
        
        if assessments_ma:
            self._fill_autoeval_ma(wb, assessments_ma, measures_additional or [])
        
        # Save to bytes
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    def _find_column_by_header(self, sheet, header_type: str, header_row: int = 1) -> Optional[int]:
        """Find column index by matching header text."""
        matchers = COLUMN_MATCHERS.get(header_type, [])
        for col_idx in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=header_row, column=col_idx).value
            if cell_value:
                cell_lower = str(cell_value).lower().strip()
                for matcher in matchers:
                    if matcher in cell_lower:
                        return col_idx
        return None
    
    def _find_sheet_by_name(self, wb, possible_names: List[str]):
        """Find sheet by trying multiple possible names."""
        for name in possible_names:
            if name in wb.sheetnames:
                return wb[name]
            # Try case-insensitive match
            for sheet_name in wb.sheetnames:
                if sheet_name.lower().strip() == name.lower().strip():
                    return wb[sheet_name]
        return None
    
    def _fill_autoeval_mg(self, wb, assessments: List[Dict[str, Any]]):
        """Fill the Autoeval MG sheet with assessment data."""
        sheet = self._find_sheet_by_name(wb, ['Autoeval MG', 'Autoevaluación MG', 'AutoevalMG'])
        if not sheet:
            return
        
        # Find header row (usually row 1 or 2)
        header_row = 1
        for row in range(1, 5):
            if sheet.cell(row=row, column=1).value:
                header_row = row
                break
        
        # Find columns
        mg_col = self._find_column_by_header(sheet, 'mg_id', header_row)
        subpart_col = self._find_column_by_header(sheet, 'subpart_id', header_row)
        difficulty_col = self._find_column_by_header(sheet, 'difficulty', header_row)
        maturity_col = self._find_column_by_header(sheet, 'maturity', header_row)
        
        if not all([difficulty_col, maturity_col]):
            # Try to detect by position (common layout)
            # Usually: MG | Apartado | Descripción | Dificultad | Madurez | Estado | Plan
            difficulty_col = difficulty_col or 4
            maturity_col = maturity_col or 5
        
        # Create lookup by (mg_id, subpart_id) for faster matching
        assessment_lookup = {}
        for a in assessments:
            key = (a.get('mg_id', ''), a.get('subpart_id', ''))
            assessment_lookup[key] = a
        
        # Iterate through data rows
        for row in range(header_row + 1, sheet.max_row + 1):
            mg_value = sheet.cell(row=row, column=mg_col or 1).value if mg_col else None
            subpart_value = sheet.cell(row=row, column=subpart_col or 2).value if subpart_col else None
            
            if mg_value and subpart_value:
                key = (str(mg_value).strip(), str(subpart_value).strip())
                assessment = assessment_lookup.get(key)
                
                if assessment:
                    if assessment.get('difficulty') and difficulty_col:
                        sheet.cell(row=row, column=difficulty_col).value = assessment['difficulty']
                    if assessment.get('maturity') and maturity_col:
                        sheet.cell(row=row, column=maturity_col).value = assessment['maturity']
    
    def _fill_measures_additional(self, wb, measures: List[Dict[str, Any]]):
        """Fill the Medidas Adicionales sheet."""
        sheet = self._find_sheet_by_name(wb, ['Medidas Adicionales', 'MA', 'Medidas adicionales'])
        if not sheet:
            return
        
        # Find header row
        header_row = 1
        for row in range(1, 5):
            if sheet.cell(row=row, column=1).value:
                header_row = row
                break
        
        # Find columns
        desc_col = self._find_column_by_header(sheet, 'description', header_row) or 2
        file_col = self._find_column_by_header(sheet, 'file_name', header_row) or 3
        
        # Write measures starting after header
        for idx, measure in enumerate(measures):
            row = header_row + 1 + idx
            
            # MA ID in first column
            sheet.cell(row=row, column=1).value = measure.get('id', f'MA_{idx+1}')
            
            # Description
            sheet.cell(row=row, column=desc_col).value = measure.get('description', '')
            
            # File name
            if measure.get('file_name'):
                sheet.cell(row=row, column=file_col).value = measure['file_name']
    
    def _fill_relation_ma(self, wb, relations: List[Dict[str, str]], measures: List[Dict[str, Any]]):
        """Fill the Relación MA-Apart sheet with X marks."""
        sheet = self._find_sheet_by_name(wb, ['Relación MA-Apart', 'Relación MA-Apartado', 'RelMA-Apart'])
        if not sheet:
            return
        
        # This is typically a matrix with MA IDs in columns and subparts in rows
        # Mark with 'X' where relationships exist
        
        # Create lookup
        relation_set = {(r['ma_id'], r['subpart_id']) for r in relations}
        
        # Find header row with MA IDs
        header_row = 1
        ma_columns = {}  # ma_id -> column index
        
        for col in range(2, sheet.max_column + 1):
            cell_value = sheet.cell(row=header_row, column=col).value
            if cell_value and str(cell_value).startswith('MA'):
                ma_columns[str(cell_value).strip()] = col
        
        # Fill X marks
        for row in range(header_row + 1, sheet.max_row + 1):
            subpart_id = sheet.cell(row=row, column=1).value
            if subpart_id:
                subpart_str = str(subpart_id).strip()
                for ma_id, col in ma_columns.items():
                    if (ma_id, subpart_str) in relation_set:
                        sheet.cell(row=row, column=col).value = 'X'
    
    def _fill_autoeval_ma(self, wb, assessments: List[Dict[str, Any]], measures: List[Dict[str, Any]]):
        """Fill the Autoeval MA sheet with assessment data."""
        sheet = self._find_sheet_by_name(wb, ['Autoeval MA', 'Autoevaluación MA', 'AutoevalMA'])
        if not sheet:
            return
        
        # Similar logic to autoeval MG
        header_row = 1
        for row in range(1, 5):
            if sheet.cell(row=row, column=1).value:
                header_row = row
                break
        
        difficulty_col = self._find_column_by_header(sheet, 'difficulty', header_row) or 4
        maturity_col = self._find_column_by_header(sheet, 'maturity', header_row) or 5
        
        # Create lookup
        assessment_lookup = {}
        for a in assessments:
            key = (a.get('ma_id', ''), a.get('subpart_id', ''))
            assessment_lookup[key] = a
        
        # Fill data
        for row in range(header_row + 1, sheet.max_row + 1):
            ma_value = sheet.cell(row=row, column=1).value
            subpart_value = sheet.cell(row=row, column=2).value
            
            if ma_value and subpart_value:
                key = (str(ma_value).strip(), str(subpart_value).strip())
                assessment = assessment_lookup.get(key)
                
                if assessment:
                    if assessment.get('difficulty'):
                        sheet.cell(row=row, column=difficulty_col).value = assessment['difficulty']
                    if assessment.get('maturity'):
                        sheet.cell(row=row, column=maturity_col).value = assessment['maturity']


def list_available_templates(templates_dir: str = None) -> Dict[str, bool]:
    """
    List all expected templates and their availability.
    
    Returns:
        Dict mapping requirement_code to boolean (True if template exists)
    """
    filler = TemplateFiller(templates_dir)
    result = {}
    for code in TEMPLATE_MAPPING:
        path = filler.get_template_path(code)
        result[code] = path is not None and path.exists()
    return result
