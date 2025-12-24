"""
Generador de Excel - Motor de Exportación (9 pestañas)
Sistema de Preevaluación Sandbox IA España
"""
from io import BytesIO
from typing import Any, Dict, List, Optional
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Colores corporativos Garrigues
GARRIGUES_GREEN = "004438"  # PANTONE 3308 C
BRIGHT_GREEN = "009A77"
WHITE = "FFFFFF"
LIGHT_GRAY = "F5F5F5"
LIGHT_GREEN = "D4EDDA"
LIGHT_YELLOW = "FFF3CD"


def create_header_style() -> Dict:
    """Estilo para encabezados de tabla."""
    return {
        'font': Font(bold=True, color=WHITE),
        'fill': PatternFill(start_color=GARRIGUES_GREEN, end_color=GARRIGUES_GREEN, fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }


def create_cell_style() -> Dict:
    """Estilo para celdas normales."""
    return {
        'alignment': Alignment(vertical='center', wrap_text=True),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }


def apply_style(cell, style_dict: Dict):
    """Aplica un diccionario de estilos a una celda."""
    for attr, value in style_dict.items():
        setattr(cell, attr, value)


def render_portada(ws, data: Dict):
    """Pestaña 1: Portada con datos del proyecto."""
    ws.merge_cells('A1:D1')
    ws['A1'] = 'INFORME DE PREEVALUACIÓN - SANDBOX IA ESPAÑA'
    ws['A1'].font = Font(size=20, bold=True, color=GARRIGUES_GREEN)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws['A3'] = 'Fecha de generación:'
    ws['B3'] = datetime.now().strftime('%d/%m/%Y %H:%M')
    
    ws['A5'] = 'DATOS DEL PROYECTO'
    ws['A5'].font = Font(size=14, bold=True, color=GARRIGUES_GREEN)
    
    row = 7
    metadata = data.get('project_metadata', {})
    fields = [
        ('Nombre del Proyecto', metadata.get('nombre', '-')),
        ('Sector', metadata.get('sector', '-')),
        ('TRL (Technology Readiness Level)', metadata.get('trl', '-')),
        ('Proveedor', metadata.get('proveedor', '-')),
        ('Descripción', metadata.get('descripcion', '-')),
    ]
    
    for label, value in fields:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = str(value) if value else '-'
        row += 1
    
    # Disclaimer
    row += 2
    ws[f'A{row}'] = 'AVISO LEGAL'
    ws[f'A{row}'].font = Font(bold=True, color='C00000')
    row += 1
    ws.merge_cells(f'A{row}:D{row+2}')
    ws[f'A{row}'] = (
        'Este informe es una herramienta de autodiagnóstico y no constituye '
        'una evaluación oficial del Sandbox de IA. Los resultados deben ser '
        'validados por profesionales cualificados antes de su presentación.'
    )
    ws[f'A{row}'].alignment = Alignment(wrap_text=True)
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 50


def render_intro(ws):
    """Pestaña 2: Introducción y escala L1-L8."""
    ws['A1'] = 'ESCALA DE MADUREZ (L1-L8)'
    ws['A1'].font = Font(size=16, bold=True, color=GARRIGUES_GREEN)
    
    ws['A3'] = 'Según la Guía 16 de AESIA, el nivel de madurez se evalúa en 8 niveles:'
    
    levels = [
        ('L1', 'No identificada', 'La medida no ha sido identificada como necesaria'),
        ('L2', 'Identificada, no documentada', 'Se conoce la necesidad pero no está documentada'),
        ('L3', 'Documentada, no implementada', 'Existe documentación pero no implementación'),
        ('L4', 'Parcialmente implementada', 'Implementación parcial sin cobertura completa'),
        ('L5', 'Implementada sin evidencia', 'Implementada pero sin evidencia documentada'),
        ('L6', 'Implementada, evidencia parcial', 'Implementada con evidencia parcial'),
        ('L7', 'Implementada, evidencia completa', 'Implementada con evidencia completa'),
        ('L8', 'Cumplimiento total verificado', 'Cumplimiento verificado y validado'),
    ]
    
    # Headers
    row = 5
    headers = ['Nivel', 'Título', 'Descripción']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        apply_style(cell, create_header_style())
    
    # Data
    for level, title, desc in levels:
        row += 1
        ws.cell(row=row, column=1, value=level)
        ws.cell(row=row, column=2, value=title)
        ws.cell(row=row, column=3, value=desc)
        for col in range(1, 4):
            apply_style(ws.cell(row=row, column=col), create_cell_style())
    
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 60


def render_requirements(ws, requirements: List[Dict]):
    """Pestaña 3: Listado de requisitos del RIA."""
    ws['A1'] = 'REQUISITOS DEL REGLAMENTO DE IA'
    ws['A1'].font = Font(size=16, bold=True, color=GARRIGUES_GREEN)
    
    headers = ['ID', 'Artículo', 'Título', 'Descripción']
    row = 3
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        apply_style(cell, create_header_style())
    
    for req in requirements:
        row += 1
        values = [req.get('id', ''), req.get('article_ref', ''), req.get('title', ''), req.get('description', '')]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=str(value) if value else '')
            apply_style(cell, create_cell_style())
    
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 60


def render_measures(ws, measures: List[Dict]):
    """Pestaña 4: Listado de Medidas Guía."""
    ws['A1'] = 'MEDIDAS GUÍA (MG)'
    ws['A1'].font = Font(size=16, bold=True, color=GARRIGUES_GREEN)
    
    ws['A3'] = 'Catálogo de medidas según las Guías AESIA para cada requisito del RIA.'
    
    headers = ['ID Medida', 'Requisito', 'Guía', 'Descripción']
    row = 5
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        apply_style(cell, create_header_style())
    
    # Medidas de ejemplo basadas en los requisitos
    sample_measures = [
        {'id': 'MG_01_01', 'req': 'REQ_01', 'guide': 'Guía 4', 'desc': 'Identificar y analizar los riesgos conocidos y previsibles'},
        {'id': 'MG_01_02', 'req': 'REQ_01', 'guide': 'Guía 4', 'desc': 'Estimar y evaluar los riesgos que puedan surgir'},
        {'id': 'MG_01_03', 'req': 'REQ_01', 'guide': 'Guía 4', 'desc': 'Evaluar otros riesgos basándose en datos de seguimiento'},
        {'id': 'MG_01_04', 'req': 'REQ_01', 'guide': 'Guía 4', 'desc': 'Adoptar medidas de gestión de riesgos adecuadas'},
        {'id': 'MG_02_01', 'req': 'REQ_02', 'guide': 'Guía 5', 'desc': 'Establecer prácticas de gobernanza de datos'},
        {'id': 'MG_02_02', 'req': 'REQ_02', 'guide': 'Guía 5', 'desc': 'Examinar posibles sesgos en los datos'},
        {'id': 'MG_02_03', 'req': 'REQ_02', 'guide': 'Guía 5', 'desc': 'Identificar lagunas o deficiencias en los datos'},
        {'id': 'MG_03_01', 'req': 'REQ_03', 'guide': 'Guía 6', 'desc': 'Preparar documentación técnica completa'},
        {'id': 'MG_03_02', 'req': 'REQ_03', 'guide': 'Guía 6', 'desc': 'Mantener documentación actualizada'},
        {'id': 'MG_09_01', 'req': 'REQ_09', 'guide': 'Guía 12', 'desc': 'Estrategia de cumplimiento regulatorio'},
        {'id': 'MG_09_02', 'req': 'REQ_09', 'guide': 'Guía 12', 'desc': 'Técnicas y procedimientos de diseño'},
        {'id': 'MG_09_03', 'req': 'REQ_09', 'guide': 'Guía 12', 'desc': 'Examen, prueba y validación'},
        {'id': 'MG_09_04', 'req': 'REQ_09', 'guide': 'Guía 12', 'desc': 'Gestión de modificaciones'},
    ]
    
    for measure in sample_measures:
        row += 1
        values = [measure['id'], measure['req'], measure['guide'], measure['desc']]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            apply_style(cell, create_cell_style())
    
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 60


def render_rel_mg(ws):
    """Pestaña 5: Matriz de relación Medidas-Requisitos."""
    ws['A1'] = 'MATRIZ DE RELACIÓN MG - REQUISITOS'
    ws['A1'].font = Font(size=16, bold=True, color=GARRIGUES_GREEN)
    
    ws['A3'] = 'Relación entre Medidas Guía y los Requisitos del RIA que cubren.'
    
    # Headers
    requirements = ['REQ_01', 'REQ_02', 'REQ_03', 'REQ_09']
    row = 5
    ws.cell(row=row, column=1, value='Medida')
    apply_style(ws.cell(row=row, column=1), create_header_style())
    
    for col, req in enumerate(requirements, 2):
        cell = ws.cell(row=row, column=col, value=req)
        apply_style(cell, create_header_style())
    
    # Matriz
    measures_map = {
        'MG_01_01': [1, 0, 0, 0],
        'MG_01_02': [1, 0, 0, 0],
        'MG_02_01': [0, 1, 0, 0],
        'MG_02_02': [0, 1, 0, 0],
        'MG_03_01': [0, 0, 1, 0],
        'MG_09_01': [0, 0, 0, 1],
        'MG_09_02': [0, 0, 0, 1],
    }
    
    for measure_id, mapping in measures_map.items():
        row += 1
        ws.cell(row=row, column=1, value=measure_id)
        apply_style(ws.cell(row=row, column=1), create_cell_style())
        
        for col, value in enumerate(mapping, 2):
            cell = ws.cell(row=row, column=col, value='✓' if value else '')
            apply_style(cell, create_cell_style())
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if value:
                cell.fill = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type='solid')
    
    ws.column_dimensions['A'].width = 12
    for i in range(2, 6):
        ws.column_dimensions[get_column_letter(i)].width = 10


def render_assessments_mg(ws, assessments: List[Dict]):
    """Pestaña 6: Autoevaluación de Medidas Guía."""
    ws['A1'] = 'AUTOEVALUACIÓN DE MEDIDAS GUÍA'
    ws['A1'].font = Font(size=16, bold=True, color=GARRIGUES_GREEN)
    
    ws['A3'] = 'Resultados de la autoevaluación según los niveles de madurez L1-L8.'
    
    headers = ['ID Medida', 'Dificultad', 'Madurez', 'Estado', 'Plan de Adaptación']
    row = 5
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        apply_style(cell, create_header_style())
    
    for assessment in assessments:
        row += 1
        status = 'Diagnosticada' if assessment.get('diagnosis_status') == '01' else 'Pendiente'
        plan_text = f"{assessment.get('adaptation_plan', '-')} - {assessment.get('adaptation_plan_desc', '')}"
        
        values = [
            assessment.get('measure_id', ''),
            assessment.get('difficulty', '-'),
            assessment.get('maturity', '-'),
            status,
            plan_text,
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=str(value))
            apply_style(cell, create_cell_style())
            
            # Color para estado
            if col == 4:
                if value == 'Diagnosticada':
                    cell.fill = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type='solid')
                else:
                    cell.fill = PatternFill(start_color=LIGHT_YELLOW, end_color=LIGHT_YELLOW, fill_type='solid')
    
    for i, width in enumerate([15, 12, 12, 15, 35], 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def render_measures_ma(ws):
    """Pestaña 7: Medidas Adicionales."""
    ws['A1'] = 'MEDIDAS ADICIONALES (MA)'
    ws['A1'].font = Font(size=16, bold=True, color=GARRIGUES_GREEN)
    
    ws['A3'] = 'Medidas adicionales propuestas por el usuario fuera del catálogo AESIA.'
    
    headers = ['ID', 'Título', 'Descripción', 'Doc. Aportada', 'Estado SEDIA', 'Comentarios']
    row = 5
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        apply_style(cell, create_header_style())
    
    # Mensaje si no hay MAs
    row += 1
    ws.cell(row=row, column=1, value='(Sin medidas adicionales definidas)')
    ws.cell(row=row, column=1).font = Font(italic=True, color='666666')
    
    for i, width in enumerate([10, 25, 40, 15, 15, 30], 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def render_rel_ma(ws):
    """Pestaña 8: Matriz de vinculación MA-Requisitos."""
    ws['A1'] = 'MATRIZ DE VINCULACIÓN MA - REQUISITOS'
    ws['A1'].font = Font(size=16, bold=True, color=GARRIGUES_GREEN)
    
    ws['A3'] = 'Relación N:M entre Medidas Adicionales y Requisitos del RIA.'
    
    headers = ['Medida Adicional', 'REQ_01', 'REQ_02', 'REQ_03', 'REQ_09']
    row = 5
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        apply_style(cell, create_header_style())
    
    row += 1
    ws.cell(row=row, column=1, value='(Sin medidas adicionales)')
    ws.cell(row=row, column=1).font = Font(italic=True, color='666666')
    
    ws.column_dimensions['A'].width = 20
    for i in range(2, 6):
        ws.column_dimensions[get_column_letter(i)].width = 10


def render_assessments_ma(ws):
    """Pestaña 9: Autoevaluación de Medidas Adicionales."""
    ws['A1'] = 'AUTOEVALUACIÓN DE MEDIDAS ADICIONALES'
    ws['A1'].font = Font(size=16, bold=True, color=GARRIGUES_GREEN)
    
    ws['A3'] = 'Evaluación de madurez para cada MA en el contexto de cada Requisito vinculado.'
    
    headers = ['Medida', 'Requisito', 'Dificultad', 'Madurez', 'Estado', 'Plan']
    row = 5
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        apply_style(cell, create_header_style())
    
    row += 1
    ws.cell(row=row, column=1, value='(Sin evaluaciones de MAs)')
    ws.cell(row=row, column=1).font = Font(italic=True, color='666666')
    
    for i, width in enumerate([15, 10, 12, 12, 15, 30], 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def generate_excel(application_data: Dict) -> bytes:
    """
    Genera el archivo Excel completo con 9 pestañas.
    
    Args:
        application_data: Diccionario con todos los datos de la aplicación
        
    Returns:
        bytes del archivo Excel
    """
    wb = Workbook()
    
    # 1. Portada
    ws_portada = wb.active
    ws_portada.title = "1. Portada"
    render_portada(ws_portada, application_data)
    
    # 2. Intro
    ws_intro = wb.create_sheet("2. Intro")
    render_intro(ws_intro)
    
    # 3. Artículo RIA
    ws_req = wb.create_sheet("3. Artículo RIA")
    render_requirements(ws_req, application_data.get('requirements', []))
    
    # 4. Medidas Guía
    ws_mg = wb.create_sheet("4. Medidas Guía")
    render_measures(ws_mg, application_data.get('measures', []))
    
    # 5. Relación MG
    ws_rel_mg = wb.create_sheet("5. Relación MG")
    render_rel_mg(ws_rel_mg)
    
    # 6. Autoev. MG
    ws_autoev_mg = wb.create_sheet("6. Autoev. MG")
    render_assessments_mg(ws_autoev_mg, application_data.get('assessments_mg', []))
    
    # 7. Medidas MA
    ws_ma = wb.create_sheet("7. Medidas MA")
    render_measures_ma(ws_ma)
    
    # 8. Relación MA
    ws_rel_ma = wb.create_sheet("8. Relación MA")
    render_rel_ma(ws_rel_ma)
    
    # 9. Autoev. MA
    ws_autoev_ma = wb.create_sheet("9. Autoev. MA")
    render_assessments_ma(ws_autoev_ma)
    
    # Guardar como bytes
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
