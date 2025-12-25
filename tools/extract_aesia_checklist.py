import argparse
import json
import re
import sys
import openpyxl
from pathlib import Path
from typing import Dict, List, Any

def normalize_text(text: str) -> str:
    """Normalize text by stripping whitespace and handling None."""
    if text is None:
        return ""
    return str(text).strip()

def parse_article_sheet(wb, sheet_name: str) -> List[Dict[str, Any]]:
    """Extracts subparts from the Article sheet."""
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found")
    
    sheet = wb[sheet_name]
    subparts = []
    
    # Find header row
    header_row_idx = None
    for i, row in enumerate(sheet.iter_rows(values_only=True), 1):
        row_str = [str(c).lower() for c in row if c]
        if any("apartado" in c for c in row_str) and any("descripción" in c for c in row_str):
            header_row_idx = i
            break
            
    if not header_row_idx:
        print(f"Warning: Could not find header row in {sheet_name}, assuming row 1 or trying best effort")
        header_row_idx = 1

    # Simple column mapping by index (robustness improvements possible)
    # Assuming columns: Apartado | Descripción | Descripción resumida
    
    for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not row or not row[0]: continue
        
        subpart_id = normalize_text(row[0])
        description = normalize_text(row[1])
        short_desc = normalize_text(row[2]) if len(row) > 2 else ""
        
        # Stop on empty rows that look like end of table
        if not subpart_id: break
        
        subparts.append({
            "subpart_id": subpart_id,
            "description": description,
            "short_description": short_desc
        })
        
    return subparts

def parse_measures_sheet(wb, sheet_name: str) -> List[Dict[str, Any]]:
    """Extracts MG details from Measures sheet."""
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found")
        
    sheet = wb[sheet_name]
    measures = []
    
    header_row_idx = None
    for i, row in enumerate(sheet.iter_rows(values_only=True), 1):
        row_str = [str(c).lower() for c in row if c]
        if any("idmedida" in c for c in row_str) and any("descripción" in c for c in row_str):
            header_row_idx = i
            break
            
    if not header_row_idx:
        header_row_idx = 1
        
    for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not row or not row[0]: continue
        
        mg_id = normalize_text(row[0])
        description = normalize_text(row[1])
        guidance = normalize_text(row[2]) if len(row) > 2 else ""
        
        if not mg_id.upper().startswith("MG"): break
        
        measures.append({
            "mg_id": mg_id,
            "description": description,
            "guidance_questions": [q.strip() for q in guidance.split('\n') if q.strip()]
        })
        
    return measures

def parse_relation_sheet(wb, sheet_name: str, valid_mgs: List[str], valid_subparts: List[str]) -> List[Dict[str, str]]:
    """Extracts MG-Subpart relations from the Matrix sheet."""
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found")
        
    sheet = wb[sheet_name]
    relations = []
    
    # 1. Detect Matrix Structure
    # Look for the row that has MG IDs in columns
    mg_col_map = {} # {col_index: mg_id}
    header_row_idx = None
    
    for i, row in enumerate(sheet.iter_rows(values_only=True), 1):
        # check if this row contains multiple valid MG IDs
        found_mgs = 0
        current_map = {}
        for col_idx, cell_value in enumerate(row):
            val = normalize_text(cell_value)
            if val in valid_mgs:
                found_mgs += 1
                current_map[col_idx] = val
        
        if found_mgs >= 3: # Heuristic: if we find at least 3 MGs, this is the header row
            header_row_idx = i
            mg_col_map = current_map
            break
            
    if not header_row_idx:
        raise ValueError("Could not detect Matrix header row with MG IDs")
        
    # 2. Iterate rows looking for Subparts
    for row_idx, row in enumerate(sheet.iter_rows(min_row=header_row_idx + 1, values_only=True), header_row_idx + 1):
        # The first column usually holds the subpart or article text
        # We need to extract the subpart_id from the text (e.g. "13.1. Diseño...") -> "13.1"
        
        cell_text = normalize_text(row[0])
        if not cell_text: continue
        
        # Try to match subpart ID at start of string
        # Regex for patterns like "13.1", "9.2.a", "AnexoIV.1"
        match = re.match(r'^([0-9]+\.[0-9]+(\.[a-z]+(\.[vix]+)?)?|Anexo[A-Z]+\.[0-9]+(\.[a-z])?)', cell_text)
        
        if not match:
             # Check if exact match in valid_subparts
             if cell_text in valid_subparts:
                 subpart_id = cell_text
             else:
                 continue
        else:
            subpart_id = match.group(1)
            
        if subpart_id not in valid_subparts:
            # Maybe the regex was too greedy or strict, try fuzzy check or skip
            # For now, skip to avoid bad data
            print(f"Skipping row {row_idx}: Subpart '{subpart_id}' not found in Article definitions.")
            continue
            
        # 3. Check for X in MG columns
        for col_idx, mg_id in mg_col_map.items():
            cell_val = normalize_text(row[col_idx])
            if cell_val.upper() == 'X':
                relations.append({
                    "mg_id": mg_id,
                    "subpart_id": subpart_id
                })
                
    return relations

def main():
    parser = argparse.ArgumentParser(description="Extract data from AESIA Guía 16 Excel Checklist")
    parser.add_argument("xlsx_file", help="Path to the .xlsx file")
    parser.add_argument("--req-code", required=True, help="Requirement Code (e.g. RISK_MGMT)")
    parser.add_argument("--output", help="Output JSON file path")
    
    args = parser.parse_args()
    
    file_path = Path(args.xlsx_file)
    if not file_path.exists():
        print(f"Error: File {file_path} not found")
        sys.exit(1)
        
    print(f"Processing {file_path.name} for {args.req_code}...")
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        # Hardcoded sheet names based on Guía spec, but customizable
        # "Artículo RIA", "Medidas guías (MG)", "Relación MG-Apart."
        
        print("Parsing Subparts...")
        subparts = parse_article_sheet(wb, "Artículo RIA")
        valid_subpart_ids = [s['subpart_id'] for s in subparts]
        print(f"Found {len(subparts)} subparts.")
        
        print("Parsing Measures...")
        measures = parse_measures_sheet(wb, "Medidas guías (MG)")
        valid_mg_ids = [m['mg_id'] for m in measures]
        print(f"Found {len(measures)} measures.")
        
        print("Parsing Relations...")
        relations = parse_relation_sheet(wb, "Relación MG-Apart.", valid_mg_ids, valid_subpart_ids)
        print(f"Found {len(relations)} relations.")
        
        result = {
            "requirement_code": args.req_code,
            "source_file": file_path.name,
            "subparts": subparts,
            "measures": measures,
            "relations": relations
        }
        
        if args.output:
            with open(args.output, 'w', encoding='utf-8') as f:
                json.dump(result, f, indent=2, ensure_ascii=False)
            print(f"Successfully saved to {args.output}")
        else:
            print(json.dumps(result, indent=2, ensure_ascii=False))
            
    except Exception as e:
        print(f"Error processing file: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
